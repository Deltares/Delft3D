from dataclasses import dataclass
from enum import Enum
from typing import Dict
from openpyxl import Workbook


class OutputType(Enum):
    """The type of output file read by the verschillentool.

    The output type determines which tolerances apply to the statistics
    found in the verschillentool Excel file. HIS files contain time series
    information, while MAP files contain spatially aggregated information.
    """

    HIS = "his"
    MAP = "map"


@dataclass
class Tolerance:
    """Contains tolerance values for a single statistic type.

    These tolerances represent the maximum allowed values for the
    corresponding statistics (max, bias, rms). If a statistic exceeds
    its tolerance, the model run is considered to have failed the
    validation criteria.

    Attributes
    ----------
    max : float
        The maximum allowed value for the `max` statistic.
    bias : float
        The maximum allowed value for the `bias` statistic.
    rms : float
        The maximum allowed value for the `rms` statistic.
    """

    max: float
    bias: float
    rms: float


class Variable:
    """Represents a physical variable in the verschillentool output.

    A variable corresponds to a quantity such as water level or flow velocity.
    Each variable has:
    - a name used in the Excel file,
    - a physical unit,
    - a tolerance for HIS output files,
    - a tolerance for MAP output files.

    Variables are registered dynamically and stored in a class-level registry.
    The factory method in `VerschillentoolOutput` automatically loads statistics
    for all registered variables.

    Parameters
    ----------
    name : str
        The prefix used in the Excel file for this variable (e.g. "sea_surface_height").
    unit : str
        The physical unit of the variable (e.g. "m" or "m/s").
    tol_his : Tolerance
        The tolerance values applicable to HIS output files.
    tol_map : Tolerance
        The tolerance values applicable to MAP output files.
    """

    _registry: Dict[str, "Variable"] = {}

    def __init__(self, name: str, unit: str,
                 tol_his: Tolerance, tol_map: Tolerance):
        self.name = name
        self.unit = unit
        self.tol_his = tol_his
        self.tol_map = tol_map
        Variable._registry[name] = self

    @classmethod
    def register(cls, name: str, unit: str,
                 tol_his: Tolerance, tol_map: Tolerance) -> "Variable":
        """Register a new variable with its HIS and MAP tolerances.

        Parameters
        ----------
        name : str
            The prefix used in the Excel file for this variable.
        unit : str
            The physical unit of the variable.
        tol_his : Tolerance
            Tolerances for HIS output files.
        tol_map : Tolerance
            Tolerances for MAP output files.

        Returns
        -------
        Variable
            The newly created variable instance.
        """
        return cls(name, unit, tol_his, tol_map)

    @classmethod
    def get(cls, name: str) -> "Variable":
        """Retrieve a registered variable by name."""
        return cls._registry[name]

    @classmethod
    def all(cls):
        """Return all registered variables."""
        return list(cls._registry.values())

    def tolerance(self, output_type: OutputType) -> Tolerance:
        """Return the tolerance object for the given output type.

        Parameters
        ----------
        output_type : OutputType
            The type of output file (HIS or MAP).

        Returns
        -------
        Tolerance
            The tolerance values corresponding to the output type.

        Raises
        ------
        ValueError
            If the output type is not supported.
        """
        match output_type:
            case OutputType.HIS:
                return self.tol_his
            case OutputType.MAP:
                return self.tol_map
            case _:
                raise ValueError(f"Unknown output type: {output_type}")

    def __repr__(self):
        return f"Variable({self.name!r})"


# Register built-in variables with tolerances
WATER_LEVEL = Variable.register(
    "sea_surface_height", "m",
    tol_his=Tolerance(max=0.01, bias=0.0001, rms=0.001),
    tol_map=Tolerance(max=0.05, bias=0.0001, rms=0.001),
)

FLOW_VELOCITY = Variable.register(
    "sea_water_speed", "m/s",
    tol_his=Tolerance(max=0.05, bias=0.0005, rms=0.005),
    tol_map=Tolerance(max=0.10, bias=0.0005, rms=0.005),
)


@dataclass
class Statistics:
    """Contains statistics of a sample.

    These statistics are computed by the verschillentool and represent
    aggregated measures of the difference between two model runs.

    Attributes
    ----------
    avg_max : float
        The average of the maximum differences.
    avg_bias : float
        The average bias between the model runs.
    avg_rms : float
        The average root-mean-square difference.
    max : float
        The maximum difference observed.
    """

    avg_max: float
    avg_bias: float
    avg_rms: float
    max: float


@dataclass
class VerschillentoolOutput:
    """Contains statistics of a model run.

    This class represents the parsed contents of a verschillentool Excel file.
    It contains statistics for all registered variables and the number of rows
    in the Statistics sheet.

    Attributes
    ----------
    output_type : OutputType
        The type of output file (HIS or MAP).
    statistics : dict[Variable, Statistics]
        A mapping from variables to their computed statistics.
    row_count : int
        The number of data rows in the Statistics sheet.
    """

    output_type: OutputType
    statistics: Dict[Variable, Statistics]
    row_count: int

    @staticmethod
    def from_verschillentool_workbook(workbook: Workbook,
                                      output_type: OutputType) -> "VerschillentoolOutput":
        """Create a `VerschillentoolOutput` from a verschillentool Excel workbook.

        This factory method reads the Averages, Statistics, and Maxima sheets
        from the Excel file and constructs a `VerschillentoolOutput` instance
        containing statistics for all registered variables.

        The method assumes the Excel file follows the structure produced by
        the verschillentool, including:
        - Sheet names ("Averages", "Statistics", "Maxima")
        - Naming conventions for variable fields (e.g. "<prefix>_max")

        Parameters
        ----------
        workbook : Workbook
            The Excel workbook produced by the verschillentool.
        output_type : OutputType
            The type of output file (HIS or MAP).

        Returns
        -------
        VerschillentoolOutput
            The parsed contents of the Excel file.

        Raises
        ------
        ValueError
            If expected fields for a registered variable are missing.
        """

        averages_sheet = workbook["Averages"]
        statistics_sheet = workbook["Statistics"]
        maxima_sheet = workbook["Maxima"]

        stats_dict = {
            str(name_cell.value).split(maxsplit=1)[0]: float(value_cell.value)
            for name_cell, value_cell in averages_sheet["A2:B7"]
        }

        first_col = maxima_sheet.min_column - 1
        last_col = maxima_sheet.max_column - 2

        maxima_dict = {
            str(maxima_sheet[row][first_col].value).split(maxsplit=1)[0]:
                float(maxima_sheet[row][last_col].value)
            for row in range(2, maxima_sheet.max_row + 1)
        }

        statistics: Dict[Variable, Statistics] = {}

        for variable in Variable.all():
            prefix = variable.name

            try:
                statistics[variable] = Statistics(
                    avg_max=stats_dict[f"{prefix}_max"],
                    avg_bias=stats_dict[f"{prefix}_bias"],
                    avg_rms=stats_dict[f"{prefix}_rms"],
                    max=maxima_dict[prefix],
                )
            except KeyError as exc:
                raise ValueError(
                    f"Failed to parse verschillentool output: Missing key {exc}"
                ) from exc

        return VerschillentoolOutput(
            output_type=output_type,
            statistics=statistics,
            row_count=statistics_sheet.max_row - 1,
        )
