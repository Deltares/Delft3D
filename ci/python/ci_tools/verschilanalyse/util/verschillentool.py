from dataclasses import dataclass
from enum import Enum
from typing import Dict
from openpyxl import Workbook


class OutputType(Enum):
    """The type of output file read by the verschillentool."""

    HIS = "his"
    MAP = "map"


@dataclass(frozen=True)
class Tolerance:
    """Tolerances for max, bias and rms statistics. Errors
    Are reported when these tolerances are exceeded.

    Attributes
    ----------
    max : float
        Tolerance of the `max` statistic.
    bias : float
        Tolerance of the `bias` statistic.
    rms : float
        Tolerance of the `rms` statistic.
    """

    max: float
    bias: float
    rms: float


@dataclass(frozen=True)
class VariableAttributes:
    name: str
    unit: str
    tolerances: Dict[OutputType, Tolerance]

@dataclass(frozen=True)
class Statistics:
    """Contains statistics of a sample.

    These statistics are computed by the verschillentool and represent
    aggregated measures of the difference between two model runs.

    Attributes
    ----------
    avg_max : float
        The average of the maximum differences.
    avg_bias : float
        The average bias between the model runs.
    avg_rms : float
        The average root-mean-square difference.
    max : float
        The maximum difference observed.
    """

    avg_max: float
    avg_bias: float
    avg_rms: float
    max: float

class VariableRegistry:
    def __init__(self):
        self._variables: Dict[str, VariableAttributes] = {}

    def register(self, 
                 name: str,
                 unit: str, 
                 tolerances: Dict[OutputType, Tolerance]) -> None:
        if name in self._variables:
            raise ValueError(f"Variable '{name}' already registered.")
        self._variables[name] = VariableAttributes(name, unit, tolerances)

    def get(self, name: str) -> VariableAttributes:
        return self._variables[name]

    def all(self):
        return list(self._variables.keys())


@dataclass(frozen=True)
class VerschillentoolOutput:
    """Contains statistics of a model run."""

    output_type: OutputType
    statistics: Dict[str, Statistics]
    row_count: int

    @staticmethod
    def from_verschillentool_workbook(
        workbook: Workbook,
        output_type: OutputType,
        variable_registry: VariableRegistry
    ) -> "VerschillentoolOutput":
        """
        Build a VerschillentoolOutput from an Excel workbook and injected registry.

        Parameters
        ----------
        workbook : Workbook
            Excel file exported by the verschillentool.
        output_type : OutputType
            HIS or MAP.
        variable_registry : VariableRegistry
            Registry containing variable definitions.

        Returns
        -------
        VerschillentoolOutput
        """

        averages_sheet = workbook["Averages"]
        statistics_sheet = workbook["Statistics"]
        maxima_sheet = workbook["Maxima"]

        stats_dict = {
            str(name_cell.value).split(maxsplit=1)[0]: float(value_cell.value)
            for name_cell, value_cell in averages_sheet["A2:B7"]
        }

        first_col = maxima_sheet.min_column - 1
        last_col = maxima_sheet.max_column - 2

        maxima_dict = {
            str(maxima_sheet[row][first_col].value).split(maxsplit=1)[0]:
                float(maxima_sheet[row][last_col].value)
            for row in range(2, maxima_sheet.max_row + 1)
        }

        stats_per_variable: Dict[str, Statistics] = {}

        for var_name in variable_registry.all():
            var = variable_registry.get(var_name)

            try:
                stats_per_variable[var_name] = Statistics(
                    avg_max=stats_dict[f"{var.name}_max"],
                    avg_bias=stats_dict[f"{var.name}_bias"],
                    avg_rms=stats_dict[f"{var.name}_rms"],
                    max=maxima_dict[var.name],
                )
            except KeyError as exc:
                raise ValueError(
                    f"Failed to parse verschillentool output: Missing key '{exc}' for variable '{var.name}'"
                ) from exc

        return VerschillentoolOutput(
            output_type=output_type,
            statistics=stats_per_variable,
            row_count=statistics_sheet.max_row - 1,
        )
    